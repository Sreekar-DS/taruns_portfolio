"""Export portfolio data from Excel worksheets to JSON for GitHub Pages.

Expected workbook path:
    data/portfolio_data.xlsx

The script converts worksheet tables into the same JSON files used by the
Google Sheets publishing workflow and the Jekyll site.
"""

from __future__ import annotations

import json
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "data" / "portfolio_data.xlsx"
OUTPUT_DIR = ROOT / "assets" / "data"

SHEETS = {
    "Profile & Opportunity": "profile.json",
    "Links": "links.json",
    "Publications": "publications.json",
    "Awards": "awards.json",
    "Currently Working On": "currently-working-on.json",
    "Projects": "projects.json",
    "Certifications": "certifications.json",
    "Professional Development": "professional-development.json",
    "Experience": "experience.json",
    "Education": "education.json",
    "Skills": "skills.json",
}

KEY_ALIASES = {
    "name": "name",
    "primary role": "primary_role",
    "secondary role": "secondary_role",
    "profile summary": "profile_summary",
    "opportunity label": "opportunity_label",
    "opportunity roles": "opportunity_roles",
    "platform": "platform",
    "label": "label",
    "url": "url",
    "external": "external",
    "title": "title",
    "type": "type",
    "category": "category",
    "status": "status",
    "start date": "start_date",
    "expected finish date": "expected_finish_date",
    "completion date": "completion_date",
    "issue date": "issue_date",
    "expiry date": "expiry_date",
    "publication date": "publication_date",
    "award date": "award_date",
    "priority": "priority",
    "skills": "skills",
    "tools": "tools",
    "opportunity_roles": "opportunity_roles",
    "short description": "short_description",
    "progress": "progress",
    "github link": "github_link",
    "demo link": "demo_link",
    "publication url": "publication_url",
    "display on home": "display_on_home",
    "display": "display",
    "featured": "featured",
    "display order": "display_order",
    "organization": "organization",
    "provider": "provider",
    "publisher": "publisher",
    "course type": "course_type",
    "credential id": "credential_id",
    "credential url": "credential_url",
}


def normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"\s+", " ", text)
    return KEY_ALIASES.get(text, re.sub(r"[^a-z0-9]+", "_", text).strip("_"))


def clean_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        return value.strip()
    return value


def to_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    return str(value).strip().lower() in {"yes", "y", "true", "1", "show", "featured"}


def split_list(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    return [item.strip() for item in re.split(r"[,;|]", str(value)) if item.strip()]


def sheet_to_records(ws) -> list[dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [normalize_header(cell) for cell in rows[0]]
    records: list[dict[str, Any]] = []

    for row in rows[1:]:
        record = {headers[i]: clean_value(row[i]) if i < len(row) else "" for i in range(len(headers)) if headers[i]}
        if not any(value not in ("", None) for value in record.values()):
            continue

        for list_key in ["skills", "tools", "highlights", "opportunity_roles"]:
            if list_key in record:
                record[list_key] = split_list(record.get(list_key))

        for key in list(record):
            if key.startswith("display_on_") or key in {"display", "featured", "external"}:
                record[key] = to_bool(record.get(key))

        if "progress" in record and record["progress"] != "":
            try:
                record["progress"] = int(float(record["progress"]))
            except (TypeError, ValueError):
                record["progress"] = 0

        if "display_order" in record and record["display_order"] != "":
            try:
                record["display_order"] = int(float(record["display_order"]))
            except (TypeError, ValueError):
                record["display_order"] = 999

        records.append(record)

    return sorted(records, key=lambda item: item.get("display_order", 999))


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    if not WORKBOOK_PATH.exists():
        print(f"Workbook not found at {WORKBOOK_PATH}. Keeping checked-in JSON files unchanged.")
        return

    workbook = load_workbook(WORKBOOK_PATH, data_only=True)

    for sheet_name, output_name in SHEETS.items():
        if sheet_name not in workbook.sheetnames:
            print(f"Sheet '{sheet_name}' not found. Skipping {output_name}.")
            continue

        records = sheet_to_records(workbook[sheet_name])
        output_path = OUTPUT_DIR / output_name
        output_path.write_text(json.dumps(records, indent=2, ensure_ascii=False), encoding="utf-8")
        print(f"Exported {len(records)} records to {output_path.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
